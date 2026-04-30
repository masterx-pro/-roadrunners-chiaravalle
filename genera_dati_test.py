import openpyxl
from datetime import date, timedelta
import random

wb = openpyxl.Workbook()

# ============================================================
# CATEGORIE
# ============================================================
ws = wb.active
ws.title = 'Categorie'
ws.append(['ID_Categoria', 'Nome', 'Fascia_Eta', 'Attiva'])
categorie = [
    ('CAT-01', 'Esordienti M', '9-10 anni', 'TRUE'),
    ('CAT-02', 'Esordienti F', '9-10 anni', 'TRUE'),
    ('CAT-03', 'Giovanissimi M', '11-12 anni', 'TRUE'),
    ('CAT-04', 'Giovanissimi F', '11-12 anni', 'TRUE'),
]
for c in categorie:
    ws.append(c)

# ============================================================
# ATLETI
# ============================================================
ws2 = wb.create_sheet('Atleti')
ws2.append(['ID_Atleta','Nome','Cognome','Data_Nascita','Codice_Fiscale','ID_Categoria','Genitore_Nome','Nome_Categoria','Genitore_Telefono','Genitore_Email','Scad_Certificato','Scad_FISR','Numero_FISR','Drive_Folder_ID','Attivo','Data_Iscrizione','Note','Numero_Gara'])

nomi_m = ['Luca','Marco','Andrea','Matteo','Giovanni','Francesco','Alessandro','Simone','Davide','Lorenzo','Riccardo','Federico','Stefano','Paolo','Nicola','Alberto','Emanuele','Gabriele','Tommaso','Daniele','Filippo','Roberto','Antonio','Pietro','Giacomo']
nomi_f = ['Sofia','Giulia','Martina','Sara','Valentina','Chiara','Federica','Alessia','Laura','Silvia','Elisa','Anna','Giorgia','Francesca','Elena','Beatrice','Irene','Claudia','Monica','Serena','Roberta','Paola','Cristina','Giovanna','Michela']
cognomi = ['Rossi','Bianchi','Ferrari','Esposito','Romano','Colombo','Ricci','Marino','Greco','Bruno','Gallo','Conti','De Luca','Mancini','Costa','Giordano','Rizzo','Lombardi','Moretti','Barbieri','Fontana','Santoro','Mariani','Rinaldi','Caruso','Ferrara','Galli','Martini','Leone','Longo']

atleti_ids = []
for i in range(1, 51):
    id_atleta = f'ATL-{i:03d}'
    atleti_ids.append(id_atleta)
    cat_idx = random.randint(0, 3)
    cat = categorie[cat_idx]
    is_m = 'M' in cat[1]
    nome = random.choice(nomi_m if is_m else nomi_f)
    cognome = random.choice(cognomi)
    anno_nascita = random.randint(2012, 2016)
    data_nascita = date(anno_nascita, random.randint(1,12), random.randint(1,28))
    genitore_nome = random.choice(nomi_m) + ' ' + cognome
    telefono = f'33{random.randint(1000000, 9999999)}'
    email = f'{nome.lower()}.{cognome.lower().replace(" ","")}@gmail.com'
    scad_cert = date(2026, random.randint(1,12), random.randint(1,28))
    scad_fisr = date(2026, 8, 31)
    numero_fisr = str(random.randint(100000, 999999))
    data_iscrizione = date(2023, random.randint(9,12), random.randint(1,28))
    numero_gara = str(random.randint(1, 200))
    ws2.append([
        id_atleta, nome, cognome, data_nascita.isoformat(),
        '', cat[0], genitore_nome, cat[1],
        telefono, email, scad_cert.isoformat(),
        scad_fisr.isoformat(), numero_fisr, '', 'TRUE',
        data_iscrizione.isoformat(), '', numero_gara
    ])

# ============================================================
# PATTINI
# ============================================================
ws3 = wb.create_sheet('Pattini')
ws3.append(['ID_Pattino','Marca','Taglia','Stato','ID_Atleta','Data_Inizio_Noleggio','Stato_Pagamento','Note'])
marche = ['Rollerblade','K2','Bauer','Roces','Powerslide','Seba','USD','Fila']
stati = ['Buono','Buono','Buono','Usurato','Rotto']
atleti_con_pattini = random.sample(atleti_ids, 35)
for i in range(1, 51):
    id_pattino = f'P-{i:03d}'
    taglia = random.randint(28, 42)
    stato = random.choice(stati)
    marca = random.choice(marche)
    if i <= 35:
        id_atleta = atleti_con_pattini[i-1]
        data_inizio = date(2024, random.randint(9,12), random.randint(1,28))
        stato_pag = random.choice(['Pagato','Da pagare'])
    else:
        id_atleta = ''
        data_inizio = ''
        stato_pag = ''
    ws3.append([id_pattino, marca, taglia, stato, id_atleta,
                data_inizio.isoformat() if data_inizio else '',
                stato_pag, ''])

# ============================================================
# RUOTE
# ============================================================
ws4 = wb.create_sheet('Ruote')
ws4.append(['ID_Set','Diametro_mm','Durezza_A','Quantita','Stato','Note'])
diametri = [72, 76, 80, 84, 90, 100, 110]
durezze = [78, 80, 82, 84, 86, 88]
stati_ruote = ['Buone','Buone','Usurate','Da sostituire']
for i in range(1, 21):
    ws4.append([
        f'R-{i:03d}',
        random.choice(diametri),
        random.choice(durezze),
        random.choice([8, 16, 24]),
        random.choice(stati_ruote),
        ''
    ])

# ============================================================
# SLOT_FISSI
# ============================================================
ws5 = wb.create_sheet('Slot_Fissi')
ws5.append(['ID_Slot','ID_Categoria','Nome_Categoria','Giorno_Settimana','Ora_Inizio','Ora_Fine','Allenatore','Attivo'])
giorni = ['Lunedì','Martedì','Giovedì']
allenatori = ['Marco Verdi','Lucia Neri','Giorgio Blu']
idx = 1
for cat in categorie:
    for giorno in giorni:
        ws5.append([
            f'SF-{idx:03d}', cat[0], cat[1], giorno,
            '18:30', '20:00', random.choice(allenatori), 'TRUE'
        ])
        idx += 1

# ============================================================
# EVENTI_SPECIALI
# ============================================================
ws6 = wb.create_sheet('Eventi_Speciali')
ws6.append(['ID_Evento','Titolo','Data_Inizio','Ora_Inizio','Data_Fine','Ora_Fine','Tipo','Luogo','ID_Categoria','Scad_Iscrizione','Scad_Pagamento','Data_Convocati','Documenti_Richiesti','Iscritti','Stato_Pagamento_Gara','Note'])
tipi = ['Gara','Trasferta','Altro']
luoghi = ['Ancona','Senigallia','Jesi','Pesaro','Macerata','Civitanova','Fabriano','Martinsicuro','Pescara','Bologna']
titoli_gara = ['Gara Regionale','Campionato Provinciale','Trofeo Primavera','Gara di Esordio','Campionato Italiano','Gara Amichevole','Trofeo Estivo','Memorial Rossi','Gran Premio','Coppa Italia']
for i in range(1, 21):
    data_inizio = date(2026, random.randint(4,8), random.randint(1,28))
    data_fine = data_inizio + timedelta(days=random.randint(0,2))
    scad_isc = data_inizio - timedelta(days=random.randint(7,14))
    scad_pag = data_inizio - timedelta(days=random.randint(3,6))
    data_conv = data_inizio - timedelta(days=random.randint(1,3))
    iscritti_gara = ','.join(random.sample(atleti_ids, random.randint(5,15)))
    ws6.append([
        f'EV-{i:03d}',
        random.choice(titoli_gara),
        data_inizio.isoformat(),
        '09:00',
        data_fine.isoformat(),
        '18:00',
        random.choice(tipi),
        random.choice(luoghi),
        random.choice(categorie)[0],
        scad_isc.isoformat(),
        scad_pag.isoformat(),
        data_conv.isoformat(),
        'Tessera FISR, Certificato medico',
        iscritti_gara,
        random.choice(['Pagato','Da pagare']),
        ''
    ])

# ============================================================
# PRESENZE
# ============================================================
ws7 = wb.create_sheet('Presenze')
ws7.append(['ID_Presenza','Tipo_Sessione','ID_Riferimento','Data','ID_Atleta','Presente'])
pres_idx = 1
for settimana in range(20):
    for giorno in ['Lunedì','Martedì','Giovedì']:
        data_pres = date(2025, 9, 1) + timedelta(weeks=settimana, days=['Lunedì','Martedì','Giovedì'].index(giorno))
        ref = f'{giorno}_{data_pres.isoformat()}'
        for id_atleta in random.sample(atleti_ids, random.randint(20,40)):
            ws7.append([
                f'PR-{pres_idx:04d}',
                'Allenamento', ref,
                data_pres.isoformat(),
                id_atleta,
                random.choice(['TRUE','TRUE','TRUE','FALSE'])
            ])
            pres_idx += 1

# ============================================================
# MODULISTICA
# ============================================================
ws8 = wb.create_sheet('Modulistica')
ws8.append(['ID_Modulo','Nome','Descrizione','Drive_File_ID','Stagione','Attivo'])
moduli = [
    ('MOD-001','Liberatoria minori','Liberatoria per atleti minorenni','','2025-2026','TRUE'),
    ('MOD-002','Modulo iscrizione','Modulo di iscrizione annuale','','2025-2026','TRUE'),
    ('MOD-003','Consenso foto','Consenso pubblicazione immagini','','2025-2026','TRUE'),
    ('MOD-004','Modulo CONI','Modulo affiliazione CONI','','2025-2026','TRUE'),
    ('MOD-005','Richiesta tessera FISR','Modulo richiesta tessera federale','','2025-2026','TRUE'),
]
for m in moduli:
    ws8.append(m)

# ============================================================
# UTENTI
# ============================================================
ws9 = wb.create_sheet('Utenti')
ws9.append(['Email','Nome','Ruolo','Attivo'])
utenti = [
    ('presidente@roadrunners.it','Mario Rossi','Presidente','TRUE'),
    ('segreteria@roadrunners.it','Laura Bianchi','Segreteria','TRUE'),
    ('allenatore1@roadrunners.it','Marco Verdi','Allenatore','TRUE'),
    ('allenatore2@roadrunners.it','Lucia Neri','Allenatore','TRUE'),
    ('masterxpro@gmail.com','Mattia Prosperi','Admin','TRUE'),
]
for u in utenti:
    ws9.append(u)

# ============================================================
# LOG
# ============================================================
ws10 = wb.create_sheet('Log')
ws10.append(['Timestamp','Utente','Azione','Entita','Dettaglio'])
azioni = [
    ('Mattia Prosperi','Nuovo','Atleta','Mario Rossi aggiunto'),
    ('Mattia Prosperi','Modifica','Atleta','Laura Bianchi - cert. aggiornato'),
    ('Mattia Prosperi','Assegnazione','Pattino','P-001 a ATL-001'),
    ('Mattia Prosperi','Nuovo','Evento','Gara Regionale 15 aprile'),
    ('Mattia Prosperi','Reset','Pagamenti noleggio','Reset trimestrale'),
]
for i in range(50):
    a = random.choice(azioni)
    data_log = date(2026, random.randint(1,3), random.randint(1,28))
    ws10.append([
        f'{data_log.isoformat()} {random.randint(8,18):02d}:{random.randint(0,59):02d}',
        a[0], a[1], a[2], a[3]
    ])

# ============================================================
# STORICO_PATTINI
# ============================================================
ws11 = wb.create_sheet('Storico_Pattini')
ws11.append(['ID_Pattino','Marca','Taglia','ID_Atleta','Nome_Atleta','Data_Inizio','Data_Fine'])
for i in range(30):
    id_pattino = f'P-{random.randint(1,50):03d}'
    marca = random.choice(marche)
    taglia = random.randint(28,42)
    id_atleta = random.choice(atleti_ids)
    nome_atleta = random.choice(nomi_m + nomi_f) + ' ' + random.choice(cognomi)
    data_inizio = date(2023, random.randint(9,12), random.randint(1,28))
    data_fine = date(2024, random.randint(1,8), random.randint(1,28))
    ws11.append([id_pattino, marca, taglia, id_atleta, nome_atleta,
                 data_inizio.isoformat(), data_fine.isoformat()])

# Salva
wb.save('C:/Road runner/dati_test_roadrunners.xlsx')
print('File Excel creato: C:/Road runner/dati_test_roadrunners.xlsx')
